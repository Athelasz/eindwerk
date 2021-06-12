from itertools import islice
import openpyxl as px
from datetime import datetime as dt
from backports.zoneinfo import ZoneInfo as zi

def parse_time(column, do, tzone_str):
    # merge date + hour into one string, choose method based on object type
    if type(do['date_meeting']) == dt:
        d_str = do['date_meeting']
        d = str(d_str.day)+'/'+str(d_str.month)+'/' + \
            str(d_str.year)+' '+do[column+'_hour'] + ':00'
    else:
        d = do['date_meeting']+' '+do[column+'_hour'] + ':00'
    # create a datetime object
    date = dt.strptime(d, '%d/%m/%Y %H:%M:%S').replace(tzinfo=zi(tzone_str))
    # make the date object timezone aware
    date.replace(tzinfo=zi(tzone_str))
    # return the date in isoformat
    return date.isoformat()


def get_headers(sheet):
    # list to hold header information
    headers = []
    # iterate through columns to fetch header and position
    for i in sheet.iter_cols():
        headers.append(i[0].value)
    return headers


def get_rows(sheet):
    # List to hold row values
    mt_list = []
    # Iterate through each row in worksheet and fetch values into tuples
    for row in islice(sheet.values, 1, sheet.max_row):
        mt_list.append(row)
    return mt_list


def get_meetings(mt_list, headers):
    # Create a set of unique meetings
    unimt_dicts = {}
    for mt in mt_list:
        if not mt[0] in unimt_dicts:
            unimt_dict = {
                headers[0]: mt[0],
                headers[1]: mt[1],
                headers[2]: str(mt[2])[0:5],
                headers[3]: str(mt[3])[0:5]
            }
            unimt_dicts[mt[0]] = unimt_dict
    return unimt_dicts


def get_invitees(uni_mt, mt_list):
    inv_dicts = {}
    for mt in uni_mt.items():
        inv_list = []
        inv_dict = {}
        # create a dict per row in the list, using the col header 
        # as key and col location as key and value index
        for i in mt_list:
            if mt[0] in i:
                inv_dict = {
                    'displayname': i[4]+' '+i[5],
                    'email': i[6],
                    'coHost': i[7]}
                inv_list.append(inv_dict.copy())
        inv_dicts[mt[0]] = inv_list
    return inv_dicts


def main(file='list_webex.xlsx',tzone_str='Europe/Brussels'):
    """**main** - Converts an excel file to a dict suitable for creating
    webex meetings via the Webex REST api with lists and dicts.

    :param file: source data do convert, defaults to 'list_webex.xlsx'
    :type file: str, optional
    :param tzone_str: desired timezone, defaults to 'Europe/Brussels'
    :type tzone_str: str, optional
    
    |

    """
    # Open the workbook and select a worksheet
    wb = px.load_workbook(file)
    sheet = wb['list_webex']
    headers = get_headers(sheet)
    rows = get_rows(sheet)
    mt_dict = get_meetings(rows, headers)
    invitees = get_invitees(mt_dict, rows)
# Create list to collect information to convert to json
    mts_tree = {}
    for mt in mt_dict.items():
        # replace mt by mt values to avoid unneeded '[1]' entries
        mt = mt[1]
        # put title in a var because it is used multiple times, readability
        title = mt['name_meeting']
        # Create dict object with mt name + aggregated members list
        mt_tree = {
            'title': title,
            'start': parse_time('start', mt, tzone_str),
            'end': parse_time('end', mt, tzone_str),
            'timezone': tzone_str,
            'enabledAutoRecordMeeting': False,
            'allowAnyUserToBeCoHost': False,
            'invitees': invitees[title]
        }
        # Add one mt dict object to final dataset
        mts_tree[title] = mt_tree
    return mts_tree


# execute main when called directly
if __name__ == '__main__':
    main()
